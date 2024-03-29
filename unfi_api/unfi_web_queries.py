from __future__ import print_function

import datetime
import json
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, timedelta
from html import escape
from typing import List

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

import unfi_api
from unfi_api import UnfiAPI, UnfiApiClient
from unfi_api.invoice import OrderList, OrderListing
from unfi_api.old_product import product_info
from unfi_api.settings import (invoice_xhr, ridgefield_cust_num,
                               ridgefield_warehouse, search_url,
                               set_default_account_xhr, user_id,
                               xdock_cust_num, xdock_warehouse)
from unfi_api.tools import Threading
from unfi_api.utils.collections import divide_chunks
from unfi_api.utils.output import auto_size_worksheet_columns, size_cols
from unfi_api.utils.string import strings_to_numbers


def run():
    pass


def uncaught_exception_handler(exc_type, value, tb):
    print(exc_type, value, tb)
    input("PROCESS FAILED PRESS ENTER TO QUIT")


# sys.excepthook = uncaught_exception_handler


def pull_invoices(token, day: date, client: UnfiApiClient = None, num_days=1):
    invoice_dict = {}
    threading = Threading()
    dock = "Ridgefield"

    invoices = get_invoice_list(day, num_days=num_days, client=client)
    print(
        "Downloading {COUNT} {DOCK} invoices...".format(DOCK=dock, COUNT=len(invoices))
    )

    def _add_invoice_to_dict(i):
        invoice_dict[i] = get_invoice(i, token, api=client.api)

    threading.thread_with_progressbar(_add_invoice_to_dict, invoices)
    # for invoice in invoices:
    #     _add_invoice_to_dict(invoice)

    # if xdock:
    #     print("Downloading XDOCK Invoices...")
    #     xdock_invoices = get_invoice_list(token, date, num_days=2, xdock=xdock)
    #     for idx, invoice in enumerate(xdock_invoices, 1):
    #         print("Pulling invoice {}/{}".format(idx, len(xdock_invoices)))
    #         invoice_dict[invoice] = get_invoice(invoice, token, xdock)
    # else:
    #     print("Downloading Ridgefield Invoices...")
    #     invoice_numbers = get_invoice_list(token, date)
    #     for idx, invoice in enumerate(invoice_numbers, 1):
    #         print("Pulling invoice {}/{}".format(idx, len(invoice_numbers)))
    #         invoice_dict[invoice] = get_invoice(invoice, token)

    return invoice_dict


def get_invoice_list(dateobj: date, num_days=1, client: UnfiApiClient = None):
    """
    get invoice list from UNFI API
    """
    invoice_list: OrderList = client.get_invoice_list()
    invoice_dict = invoice_list.orders_by_date()
    invoice_listings: List[OrderListing] = invoice_dict[
        get_most_recent_date(invoice_dict.keys(), dateobj)
    ]
    invoice_nums: List[str] = [listing.invoice_number for listing in invoice_listings]
    return invoice_nums


def get_most_recent_date(items, dateobj: date):
    found_date = None
    for day in items:
        if (dateobj - day).days < 0:
            continue
        if found_date is None or found_date < day:
            found_date = day
    return found_date


def get_invoice(invoice_number, token, callback=None, api: unfi_api.UnfiAPI = None):

    invoice_url = invoice_xhr.format(
        invoicenum=invoice_number.strip(), custnum=api.account
    )
    header = {"authorization": "{token}".format(token=token)}
    invoice = api.get(invoice_url)
    # invoice = api.order_management.order_history.get_invoice(invoice_number)
    # invoicesoup = BeautifulSoup(invoice['data'], "html.parser")
    invoicesoup = BeautifulSoup(invoice.text, "html.parser")
    tablerows = []
    eof = False
    tables = invoicesoup.findAll("table")
    for tablenum, table in enumerate(tables[5:]):
        if eof:
            break
        rows = table.findAll("tr")
        for rownum, row in enumerate(rows):
            cells = []
            webcells = row.findAll("td")
            for cell in webcells:
                cell = (
                    cell.text.replace("\\r", "")
                    .replace("\\t", "")
                    .replace("\\n", "")
                    .strip()
                )

                cell = cell
                if "Freight Details" in cell:
                    eof = True
                    break

                v = cell
                cells.append(v)
            if eof:
                break
            tablerows.append(strings_to_numbers(cells))
    return tablerows


def create_invoice_workbook(invoicedict, outputpath=None, invdate=None):
    """
    Compile a dict of invoices numbers and rows into an xlsx workbook.
    save if requested (makes re-running the report much faster)
    :param invoicedict: {invoicenum: wbrows}
    :param outputpath: Where to save invoices.xlsx if defined.
    :return:
    """
    wb = Workbook()
    defaultws = wb.active
    i = 1
    combined = []
    if not invdate:
        invdate = datetime.date.today()
    for invoice, rows in sorted(invoicedict.items()):
        if i == 1:
            header = rows[2].copy()
            header.insert(0, "Invoice #")
            header.insert(1, "Invoice Date")
            combined.append(header)
            ws = wb.active
            ws.title = invoice
            i += 1
        else:
            ws = wb.create_sheet(invoice)
        for row in rows:
            ws.append(strings_to_numbers(row))
        auto_size_worksheet_columns(ws)
        for l in rows[3:]:
            l.insert(0, invoice[:-4])
            l.insert(1, invdate.strftime("%m/%d/%y"))
            combined.append(l)

    combws = wb.create_sheet("All Invoices", 0)
    for l in combined:
        combws.append(l)

    sheets = {k: v for v, k in enumerate(wb.worksheets)}
    # sheets = sort_dict(sheets)

    workbook = []
    for ws in wb.worksheets:
        rowlist = []
        for row in ws.rows:
            rowlist.append([cell.value for cell in row])
        workbook.append(rowlist)

    if outputpath:
        wb.save(outputpath + r"\invoices.xlsx")

    return dict(sheetindex=sheets, workbook=workbook, original=wb, combined=combined)


def set_default_account(userid, account, client: UnfiApiClient, region="West"):
    token = client.auth_token
    headers = {"authorization": token}
    payload = [{"UserId": userid, "SelectedAccount": account, "SelectedRegion": region}]
    request = requests.post(set_default_account_xhr, headers=headers, json=payload)


def make_query_list(query):
    # split query by white space and remove non text
    import re

    query_list = re.split(r"\s", query)
    query_list = [re.sub(r"\W+", "", x) for x in query_list]
    try:
        query_list.remove("")
    except ValueError:
        pass
    return query_list


def run_query(client, query_list, token, xdock=False, api: UnfiAPI=None):
    """
    Search unfi web site for queries
    :type xdock: bool
    :param xdock: whether or not to search cross dock

    :param api: authorization token
    :type api: `unfi_api.api.api.UnfiAPI`
    :type query_list: list
    :param query_list: List of upc to search for
    """
    search_results = []
    num_queried = 0
    query_list = query_list.copy()
    try:
        query_list.remove("")
    except ValueError:
        pass

    query_len = len(query_list)
    remaining_terms = query_len
    query_chunks = list(divide_chunks(query_list, 120))
    product_list = []
    warehouse = "Ridgefield"
    print(
        "Running query for %d terms from %s. Spawning %s threads."
        % (query_len, warehouse, len(query_chunks))
    )

    def _search_chunk(c):

        return search_products(c, token, xdock, api)

    with ThreadPoolExecutor(max_workers=len(query_chunks)) as e:
        futures = []
        for chunk in query_chunks:
            join_query = " ".join(str(x) for x in chunk)
            print("Searching for {} terms.".format(len(chunk)))
            futures.append(e.submit(_search_chunk, join_query))

        for r in as_completed(futures):
            search_result = r.result()
            product_list.extend(search_result.get("TopProducts"))
            search_results.append(search_result)
            print("Found {} of {} items".format(len(product_list), query_len))
            remaining_terms -= len(chunk)

    print("Found {} of {} items".format(len(product_list), query_len))
    if product_list:
        products = product_info(product_list, token, xdock=xdock, api=api)
        return products
    else:
        print("No results for your search.")
        return None


def search_products(query, token, xdock=False, api=None):
    """
    Poll the UNFI API for a query string. Returns the json response.
    :param query:
    :param token:
    :param xdock:
    :return:
    """
    header = {
        "authorization": "{token}".format(
            token=token,
        )
    }
    query = escape(query)
    if xdock:
        query_url = search_url.format(
            query=query,
            userid=user_id,
            custnum=xdock_cust_num,
            warehouse=xdock_warehouse,
        )
    else:
        query_url = search_url.format(
            query=query,
            userid=user_id,
            custnum=ridgefield_cust_num,
            warehouse=ridgefield_warehouse,
        )
    # print("Searching Database....")
    response = requests.get(query_url, headers=header)
    results = json.loads(response.content)
    print("Found: ", results["TotalHits"], " results.")
    return results
