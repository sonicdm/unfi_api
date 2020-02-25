from __future__ import print_function

try:
    from past.builtins import unicode
except ImportError:
    pass
import datetime
import os

from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter

from catalogboss.formatter import size_cols
from catalogboss.catalog.products import Product
from catalogboss.catalogio import read_workbook
from catalogboss.utils import \
    strings_to_numbers, find_most_common_member, sort_dict
from catalogboss.utils.db import AbbrRepl
from catalogboss.utils import index_header, simple_round_retail
from catalogboss.utils.upc import stripcheckdigit, isnumber
from unfi_api.unfi_web_queries import create_invoice_workbook

YELLOW = 'FFFF00'
ORANGE = 'FFA500'
ABBR = AbbrRepl()

COLUMNS = {
    1: 'Invoice Number',
    2: 'Ordered',
    3: 'Shipped',
    4: 'Received Units',
    5: 'Received Cases',
    6: 'Department',
    7: 'Brand',
    8: 'Type',
    9: 'Detail',
    10: 'UPC',
    11: 'Retail',
    12: 'SRP',
    13: 'Cost',
    14: 'Prod ID',
    15: 'Our Size',
    16: 'Invoice Pack',
    17: 'Category',
}

COLS_BY_NAME = {v: k for k, v in COLUMNS.items()}
abbr = AbbrRepl()

DEPARTMENTS = {
    11: "grocery",
    12: "non-food",
    13: "dairy",
    14: "frozen",
    15: "bulk",
    41: "deli",
    42: "cheese",
    43: "bakery",
    51: "supplements",
    52: "body care",
    53: "mercentile"
}
DEPT_NAMES = {v: k for k, v in DEPARTMENTS.items()}


def go():
    pass


def parse_invoices(invwb, inventory=None):
    invoices = {}
    for ws, idx in sorted(invwb['sheetindex'].items(), key=lambda k: k[0].title):
        if ws.title == "All Invoices":
            continue
        invoice = Invoice(invwb['workbook'][idx], ws.title, inventory=inventory)
        invoices[ws.title] = invoice

    return invoices


def write_workbook(invoiceobj):
    from openpyxl import Workbook

    inventory = invoiceobj.inventory
    invoice = invoiceobj.invoice
    department = invoiceobj.department
    invoice_num = invoiceobj.invoice_number
    date = invoiceobj.date
    wbdir = r'F:\Grocery\Receiving\UNFI'
    date_str = date.strftime('%Y-%m-%d')
    filename = "{} - {}.xlsx".format(department, invoice_num)
    outputfolder = os.path.join(wbdir, date_str)
    if not os.path.exists(outputfolder):
        os.makedirs(outputfolder)
    output = os.path.join(outputfolder, filename)
    wb = Workbook()
    invoicews = wb.active
    invoicews.title = "Invoice"
    inventoryws = wb.create_sheet('Inventory')
    form_worksheet(inventoryws, inventory)
    form_worksheet(invoicews, invoice)
    wb.save(output)


def form_worksheet(ws, products):
    """
    Write rows to a worksheet from a dict of products. based off of their X,Y values.
    :param ws: 
    :param products: 
    :return: 
    """

    for y, row in products.items():
        for x, val in row['cells'].items():
            col = get_column_letter(x)
            ws["{}{}".format(col, y)] = val

    return ws


def make_header(idx):
    """
    Write the header as defined in the COLUMNS dict
    :param idx: 
    :return: 
    """
    cells = {}
    for k, v in COLUMNS.items():
        cells[k] = v
    header = make_row(idx, cells, None)
    return header


def build_cells(product):
    cells = {
        COLS_BY_NAME['Invoice Number']: getattr(product, 'invoice_number', None),
        COLS_BY_NAME['Ordered']: getattr(product, 'ordered', None),
        COLS_BY_NAME['Shipped']: getattr(product, 'shipped', None),
        COLS_BY_NAME['Received Units']: getattr(product, 'received_units', None),
        COLS_BY_NAME['Received Cases']: getattr(product, 'received_cases', None),
        COLS_BY_NAME['Department']: getattr(product, 'department', None),
        COLS_BY_NAME['Brand']: getattr(product, 'brand', None),
        COLS_BY_NAME['Type']: getattr(product, 'type', None),
        COLS_BY_NAME['Detail']: getattr(product, 'detail', None),
        COLS_BY_NAME['UPC']: getattr(product, 'upc', None),
        COLS_BY_NAME['Retail']: getattr(product, 'retail', None),
        COLS_BY_NAME['SRP']: getattr(product, 'srp', None),
        COLS_BY_NAME['Cost']: getattr(product, 'cost', None),
        COLS_BY_NAME['Prod ID']: getattr(product, 'prod_id', None),
        COLS_BY_NAME['Our Size']: getattr(product, 'size', None),
        COLS_BY_NAME['Invoice Pack']: getattr(product, 'invoice_pack', None),
        COLS_BY_NAME['Category']: getattr(product, 'category', None),
    }
    return cells


def make_row(idx, cells, product):
    row = {'idx': idx, 'cells': cells, 'product': product}
    return row


def make_filename(date, invoicenumber=None, department=None):
    if isinstance(date, datetime.datetime):
        datestr = date.strftime('%Y-%m-%d')
    elif isinstance(date, (str, unicode)):
        datestr = date
    else:
        datestr = datetime.datetime.today().strftime('%Y-%m-%d')

    invoice = " - {}".format(invoicenumber) if invoicenumber else ""
    filename = "{} - {}{}".format(datestr, department, invoice)
    return filename


class Inventory:
    def __init__(self, inventory_wb, batch_number=None, department=None):
        self.inventory_wb = inventory_wb
        self.batch_number = batch_number
        self.department = department
        self.inventory_products = {}
        self.rows = {}
        self.parse_inventory()

    def parse_inventory(self):
        ws = self.inventory_wb
        colindex = index_header(ws, 0)
        i = 1
        header = make_header(i)
        self.rows[i] = header
        for idx, row in enumerate(ws[1:]):
            i += 1
            row = strings_to_numbers(row)
            upc = unicode(row[colindex['upc code'][0]]).replace('-', '').strip()
            upc = strings_to_numbers(upc)
            retail = row[colindex['retail'][0]]
            cases = float(row[colindex['cases'][0]])
            units = float(row[colindex['units'][0]])
            size = row[colindex['unit size'][0]]
            pack = float(row[colindex['pack'][0]])
            desc = row[colindex['description'][0]]
            if not isnumber(upc):
                continue
            product = Product(upc)
            if not product.query_result:
                if "Item not found" not in desc:
                    product.detail = desc
                product.retail = float(retail)
                product.size = size

            product.received_cases += cases
            product.received_units += units
            product.inventory_pack = pack
            self.inventory_products[upc] = product
            cells = build_cells(product)
            row = make_row(i, cells, product)
            self.rows[i] = row


class Invoice(object):
    """
    UNFI Invoice Worksheet Object
    """

    def __init__(self, invoice_sheet, invoice_number=None, department=None, date=None, inventory=None):
        """
        :param invoice_sheet: list of worksheet rows 
        :param invoice_number: UNFI invoice number
        :param department: override the parsed department name
        :param date: 
        """
        self.invoice_number = invoice_number
        self._department = None
        self.department = department
        self._departments = []
        self.invoice_sheet = invoice_sheet
        self.column_fields = None
        self.invoice_products = {}
        self.date = date
        self.invoice = None
        self.rows = {}
        self.inventory = inventory if inventory else None
        self.parse_invoice()
        if not department:
            self.department = find_most_common_member(self._departments)

    @property
    def department(self):
        return self._department

    @department.setter
    def department(self, dept):
        try:
            out = int(dept)
            self._department = out
        except (ValueError, TypeError):
            self._department = DEPT_NAMES.get(dept)

    def get_invoice_indexes(self, ws, header_row):
        colindex = index_header(ws, header_row)
        indexes = dict(
            upc=colindex['upc'][0],
            srp=colindex['reg srp'][0],
            size=colindex['eaches - pack'][0],
            ordered=colindex['ord'][0],
            shipped=colindex['shp'][0],
            desc=colindex['product desc'][0],
            cost=colindex['whls. cs. t'][0],
            prod_id=colindex['product code'][0],
        )
        return colindex, indexes

    def find_invoice_header(self, ws):
        for idx, row in enumerate(ws):
            if "UPC" in row:
                return idx
        return None

    def parse_invoice(self):
        """
        The guts of the invoice scraper.
        :return: 
        """
        ws = self.invoice_sheet
        header_row = self.find_invoice_header(ws)
        colindex, indexes = self.get_invoice_indexes(ws, header_row)
        upcidx = indexes.get('upc')
        prodididx = indexes.get('prod_id')
        srpidx = indexes.get('srp')
        sizeidx = indexes.get('size')
        orderedidx = indexes.get('ordered')
        shippedidx = indexes.get('shipped')
        descidx = indexes.get('desc')
        invpackidx = indexes.get('size')
        costidx = indexes.get('cost')
        try:
            assert not any(idx is None for idx in [upcidx, srpidx, sizeidx, orderedidx, shippedidx, descidx])
        except AssertionError as e:
            print(e)
            raise AssertionError
        i = 1
        header = make_header(i)
        self.rows[i] = header
        for idx, row in enumerate(ws[header_row + 1:]):
            upc = row[upcidx]
            if upc:
                i += 1
                upc = int(stripcheckdigit(upc))
                prod_id = row[prodididx]
                invoice_pack = row[sizeidx]

                shipped = row[shippedidx]
                ordered = row[orderedidx]
                detail = ABBR.abbr_repl(row[descidx])
                if self.invoice_products.get(upc):
                    product = self.invoice_products.get(upc)
                elif self.inventory.inventory_products.get(upc):
                    product = self.inventory.inventory_products.get(upc)
                else:
                    product = Product(upc)
                product.invoice_pack = row[invpackidx]
                # product.srp = round_retails(srp)
                srp = row[srpidx]
                if isnumber(srp):
                    srp = float(row[srpidx])
                    product.srp = simple_round_retail(srp)
                cost = float(row[costidx]) if isnumber(row[srpidx]) else None
                product.ordered += int(ordered)
                product.shipped += int(shipped)
                product.invoice_number = self.invoice_number
                setattr(product, 'cost', cost)
                setattr(product, 'prod_id', prod_id)
                setattr(product, 'invoice_pack', invoice_pack)

                if product.retail and product.srp:
                    if product.retail != product.srp:
                        product.yellow = True

                if not product.detail:
                    product.detail = detail
                self.invoice_products[upc] = product
                if product.department:
                    self._departments.append(product.department)
                cells = build_cells(product)
                # Turn info into a row dict. This will all be better with excel wrapper. Perhaps this is a beta test.
                row = make_row(i, cells, product)
                self.rows[i] = row


class InventoryReport(object):
    """
    Collection of inventory reports
    """

    def __init__(self, inventorypath=None, invoicepath=None, date=None, batch=None,
                 department=None, web_result=None, outputpath=None, invoice_date=None):
        self.inventorypath = inventorypath
        self.invoicepath = invoicepath
        self.outputpath = outputpath
        self.report_date = date

        if web_result:
            self.invoiceworkbook = create_invoice_workbook(web_result, self.outputpath, invoice_date)
            if not inventorypath:
                return
        else:
            self.invoiceworkbook = read_workbook(invoicepath)

        self.inventoryworkbook = read_workbook(inventorypath)
        self.inventory = Inventory(self.inventoryworkbook, batch, department)

        self.workbook = self.invoiceworkbook.get('original')
        self.invoices = parse_invoices(self.invoiceworkbook, self.inventory)
        self.products = self.inventory.inventory_products if self.inventory else {}
        self._combine_invoice_products(self.invoices)
        self.output_rows = {}
        self._make_new_rows()

    def _combine_invoice_products(self, invoices):
        for num, invoice in sort_dict(invoices).items():
            for upc, product in invoice.invoice_products.items():
                report_product = self.products.get(upc)
                if report_product:
                    if report_product.invoice_number not in product.invoice_number:
                        report_product.invoice_number += "\n" + product.invoice_number
                        self.products[upc] = report_product

                else:
                    self.products[upc] = product

        for upc, prod in self.products.items():
            shipped = 0
            ordered = 0
            invoices = prod.invoice_number
            if invoices:
                for invoicenum in invoices.split('\n'):
                    if invoicenum:
                        item = self.invoices[invoicenum].invoice_products[upc]
                        shipped += item.shipped
                        ordered += item.ordered
                prod.ordered = ordered
                prod.shipped = shipped

    def _make_new_rows(self):

        i = 1
        header = make_header(i)
        self.output_rows[i] = header
        i += 1
        for upc, product in sort_dict(self.products).items():
            cells = build_cells(product)
            row = make_row(i, cells, product)
            self.output_rows[i] = row
            i += 1

    def write(self, outputpath):
        i = 0
        self._make_page(self.output_rows, 'Combined Invoices', i)
        i += 1
        price_rows = self._make_price_changes()
        if len(price_rows) > 1:
            self._make_page(price_rows, 'Price Changes', i)
            i += 1
        self._make_inventory_page(i)
        self.workbook.active = 0
        self.workbook.save(filename=outputpath)

    def _make_page(self, rows, title, idx=0):
        ws = self.workbook.create_sheet(title, idx)
        for y, row in rows.items():
            product = row.get('product')
            cells = row.get('cells')
            if not product:
                if y == 1:
                    ws.append(strings_to_numbers(list(cells.values())))
                continue

            for x, cell in cells.items():
                cell = strings_to_numbers(cell)
                cell_address = "{col}{row}".format(col=get_column_letter(x), row=y)
                newcell = ws[cell_address]
                newcell.value = cell
        ws.set_printer_settings(1, 'landscape')
        self._style_ws(ws)

    def _make_inventory_page(self, idx=1):
        self.workbook.create_sheet('Inventory', idx)
        for idx, row in enumerate(self.inventoryworkbook):
            pass

    def _make_price_changes(self):
        rows = {}
        i = 1
        header = make_header(i)
        rows[i] = header
        i += 1
        for upc, product in self.products.items():
            if product.srp and product.retail:
                retail = float(product.retail)
                srp = float(product.srp)
                if srp != retail:
                    if abs((retail - srp) / srp) > .05 and product.department in [11, 12, 13, 14, 15]:
                        cells = build_cells(product)
                        rows[i] = make_row(i, cells, product)
                        i += 1
        return rows

    def _style_ws(self, ws):

        rounding_check_format = [u'AND(VALUE(RIGHT(G1,1))<>5,VALUE(RIGHT(G1,1))<>9)']
        maxrow = ws.max_row

        # Price Change Check
        yellow_fill = PatternFill('solid', start_color='ffff00', end_color='ffff00')
        srp_idx = get_column_letter(COLS_BY_NAME['SRP'])
        retail_idx = get_column_letter(COLS_BY_NAME['Retail'])
        srp_rule = CellIsRule(operator='notEqual', formula=['{}1'.format(retail_idx)], fill=yellow_fill)
        retail_rule = CellIsRule(operator='notEqual', formula=['{}1'.format(srp_idx)], fill=yellow_fill)
        self._apply_rule_to_col(ws, srp_idx, srp_rule, maxrow)
        self._apply_rule_to_col(ws, retail_idx, retail_rule, maxrow)

        # Shipping Error Check
        orange_fill = PatternFill(fill_type='solid', start_color='FFA500', end_color='FFA500')
        shipped_cases_idx = get_column_letter(COLS_BY_NAME['Shipped'])
        received_cases_idx = get_column_letter(COLS_BY_NAME['Received Cases'])
        shipped_formula = ['and({shipped}1>0,{shipped}1<>{received}1)'.format(
            shipped=shipped_cases_idx,
            received=received_cases_idx
        )]
        shipped_rule = FormulaRule(formula=shipped_formula, stopIfTrue=False, fill=orange_fill)

        self._apply_rule_to_col(ws, shipped_cases_idx, shipped_rule, maxrow)
        self._apply_rule_to_col(ws, received_cases_idx, shipped_rule, maxrow)

        # Apply AutoFilter
        maxcol = get_column_letter(ws.max_column)
        maxrow = ws.max_row
        ws.auto_filter.ref = "A1:{}{}".format(maxcol, maxrow)

        # UPC Format
        upc_number_format = "000000-00000"
        upc_idx = get_column_letter(COLS_BY_NAME['UPC'])
        for cell in ws['{}:{}'.format(upc_idx, upc_idx)]:
            cell.number_format = upc_number_format
        # word wrap the invoice column
        invoices_idx = get_column_letter(COLS_BY_NAME['Invoice Number'])
        for cell in ws['{0}:{0}'.format(invoices_idx)]:
            wrap = Alignment(wrap_text=True)
            cell.alignment = wrap

        # Auto-Size the columns
        size_cols(ws)

    def _apply_rule_to_col(self, ws, col, rule, maxrow=None):
        if not maxrow:
            maxrow = ws.max_row
        ws.conditional_formatting.add('{0}1:{0}{1}'.format(col, maxrow), rule)


def pull_inventory(batchnum):
    """
    Hacky function to streamline pulling the inventory report.
    :param batchnum:
    :return:
    """
    if batchnum:
        command = r'"e:\brdata workstation\plbclient.exe" mtiw9404 -POST -BATCH={} -EXCLIND'.format(batchnum)
    else:
        command = r'"e:\brdata workstation\plbclient.exe" mtiw9404 -POST -EXCLIND'
    run_brdata = os.system(command)
    return run_brdata


if __name__ == "__main__":
    go()
